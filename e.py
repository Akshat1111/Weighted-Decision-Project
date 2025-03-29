from openpyxl import load_workbook

# Load the workbook
file_path = "example_updated_3.xlsx"  # Replace with your actual file
wb = load_workbook(file_path)
data_sheet = wb["Data"]
sheet1 = wb["Sheet1"]

# Step 1: Extract precomputed data from "Sheet1"

# Yearly average mapping (from I5:I12 → J5:J12)
year_avg_map = {}
for i in range(5, 13):
    year = sheet1.cell(row=i, column=9).value  # Year from I column
    avg_value = sheet1.cell(row=i, column=10).value  # Avg from J column
    if year is not None:
        year_avg_map[year] = avg_value

# State-level mapping (M5:O72)
state_avg_map = {}
for i in range(5, 73):
    state = sheet1.cell(row=i, column=13).value  # State from M column
    occurrences = sheet1.cell(row=i, column=14).value  # Occurrences from N column
    avg_value = sheet1.cell(row=i, column=15).value  # Average from O column
    if state is not None:
        state_avg_map[state] = (occurrences, avg_value)

# Region-level mapping (Q5:S8)
region_avg_map = {}
for i in range(5, 9):
    region = sheet1.cell(row=i, column=17).value  # Region from Q column
    occurrences = sheet1.cell(row=i, column=18).value  # Occurrences from R column
    avg_value = sheet1.cell(row=i, column=19).value  # Average from S column
    if region is not None:
        region_avg_map[region] = (occurrences, avg_value)

# Step 2: Process each column in "Data" for years (E5:L167)
for col in range(5, 13):  # E = 5, L = 12 (Columns E to L)
    year = data_sheet.cell(row=4, column=col).value  # Year header in row 4

    if year not in year_avg_map:
        continue  # Skip if the year is not found

    year_avg = year_avg_map[year]

    for row in range(5, 168):  # E5:L167
        cell = data_sheet.cell(row=row, column=col)

        # Check if the cell is actually empty
        if cell.value is not None and str(cell.value).strip() != "":
            continue  # Skip non-empty cells

        # Get state and region for this row
        state = data_sheet.cell(row=row, column=3).value  # Column C
        region = data_sheet.cell(row=row, column=4).value  # Column D

        # Get precomputed state & region averages
        state_data = state_avg_map.get(state, (0, year_avg))  # (Occurrence, Average)
        region_data = region_avg_map.get(region, (0, year_avg))  # (Occurrence, Average)

        state_occurrence, state_avg = state_data
        region_occurrence, region_avg = region_data

        # Determine weights based on state occurrence
        if state_occurrence < 3:
            weight_state = 0.0
            weight_region = 0.5
        else:
            weight_state = 0.2
            weight_region = 0.3
        weight_year = 0.5

        # Compute weighted value
        weighted_value = (weight_year * year_avg +
                          weight_region * region_avg +
                          weight_state * state_avg)

        # Print debug info
        print(f"Row {row}, Col {col} (Year {year}) - Filling: {weighted_value:.2f} "
              f"(YearAvg: {year_avg}, StateAvg: {state_avg}, RegionAvg: {region_avg})")

        # Fill the empty cell
        cell.value = weighted_value

# Save the updated file properly
wb.save("example_updated_5.xlsx")
print("✅ File saved successfully as example_updated.xlsx")
