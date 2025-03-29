from openpyxl import load_workbook
# Calculates the number of times a state has occured and make a table in excel for it.
# Load the workbook and sheets
wb = load_workbook("example_updated.xlsx")  # Replace with your file name
data_sheet = wb["Data"]
summary_sheet = wb["Sheet1"]

# Dictionary to store state information:
# Key: state string, Value: dict with 'count' and 'values' list.
state_data = {}

# Loop over rows 5 to 167 in "Data"
for row in range(5, 168):
    state = data_sheet.cell(row=row, column=3).value  # Column C (state) is index 3
    if state is None or str(state).strip() == "":
        continue  # Skip rows with empty state
    # Initialize dictionary entry if not already present
    if state not in state_data:
        state_data[state] = {'count': 0, 'values': []}
    
    # Increase the occurrence count for the state
    state_data[state]['count'] += 1

    # Loop over columns E to L (columns 5 to 12) for the current row.
    for col in range(5, 13):
        cell_value = data_sheet.cell(row=row, column=col).value
        # Check if the cell_value is numeric
        if isinstance(cell_value, (int, float)):
            state_data[state]['values'].append(cell_value)

# Create a sorted list of states for consistent output
sorted_states = sorted(state_data.keys())

# Write the headers in "sheet1" at M4, N4, and O4
summary_sheet["M4"] = "State"
summary_sheet["N4"] = "Occurrence"
summary_sheet["O4"] = "Average"

# Starting row for the summary table on "sheet1" (M5 onward)
start_row = 5

# Write each state's data to the summary table
for idx, state in enumerate(sorted_states):
    row = start_row + idx
    count = state_data[state]['count']
    values = state_data[state]['values']
    # Compute the average if there are any values; else leave it as None
    avg = sum(values) / len(values) if values else None

    summary_sheet.cell(row=row, column=13, value=state)  # Column M (13)
    summary_sheet.cell(row=row, column=14, value=count)    # Column N (14)
    summary_sheet.cell(row=row, column=15, value=avg)        # Column O (15)

# Save the updated workbook (change the filename as needed)
wb.save("example_updated_2.xlsx")
