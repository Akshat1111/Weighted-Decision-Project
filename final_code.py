from openpyxl import load_workbook
# Main code to fill data points also logs the values !
# Load the workbook
wb = load_workbook("example_updated_3.xlsx")  # Replace with your actual filename
data_sheet = wb["Data"]
sheet1 = wb["Sheet1"]

# Step 1: Extract precomputed data from "Sheet1"
# Years mapping from Sheet1 I5:I12 (year labels) → J5:J12 (corresponding avg values)
year_avg_map = {}
for i in range(5, 13):
    year_key = sheet1.cell(row=i, column=9).value  # Year from I column
    avg_value = sheet1.cell(row=i, column=10).value  # Avg from J column
    if year_key is not None:
        try:
            # Convert to int so that keys are numbers
            year_avg_map[int(year_key)] = avg_value
        except:
            year_avg_map[year_key] = avg_value

# State-level mapping from M5:O72 (State, Occurrence, Average)
state_avg_map = {sheet1.cell(row=i, column=13).value: (sheet1.cell(row=i, column=14).value,
                                                          sheet1.cell(row=i, column=15).value)
                 for i in range(5, 73) if sheet1.cell(row=i, column=13).value is not None}

# Region-level mapping from Q5:S8 (Region, Occurrence, Average)
region_avg_map = {sheet1.cell(row=i, column=17).value: (sheet1.cell(row=i, column=18).value,
                                                          sheet1.cell(row=i, column=19).value)
                  for i in range(5, 9) if sheet1.cell(row=i, column=17).value is not None}

# Debug: Print the mappings so you can verify the keys
print("Year Mapping:", year_avg_map)
print("State Mapping:", state_avg_map)
print("Region Mapping:", region_avg_map)

# Step 2: Process each column in "Data(2)" for years (E5:L167)
for col in range(5, 13):  # Columns E to L (E=5, L=12)
    year_header = data_sheet.cell(row=4, column=col).value  # Year header in row 4
    try:
        year = int(year_header)
    except:
        year = year_header

    print(f"Processing column {col}, Year: {year}")

    if year not in year_avg_map:
        print(f"Year {year} not found in year_avg_map. Skipping column {col}.")
        continue  # Skip if the year is not found

    year_avg = year_avg_map[year]

    for row in range(5, 168):  # Rows 5 to 167
        cell = data_sheet.cell(row=row, column=col)

        # Check if the cell is actually empty (ignoring spaces)
        if cell.value is not None and str(cell.value).strip() != "":
            continue  # Skip non-empty cells

        # Get state (from column C) and region (from column D) for this row
        state = data_sheet.cell(row=row, column=3).value  # Column C
        region = data_sheet.cell(row=row, column=4).value  # Column D

        # Get precomputed state & region averages
        state_data = state_avg_map.get(state, (0, year_avg))  # (Occurrence, Average)
        region_data = region_avg_map.get(region, (0, year_avg))  # (Occurrence, Average)

        state_occurrence, state_avg = state_data
        region_occurrence, region_avg = region_data

        # Determine weights based on state occurrence
        if state_occurrence < 3:
            weight_state = 0.0
            weight_region = 0.5
        else:
            weight_state = 0.2
            weight_region = 0.3
        weight_year = 0.5

        # Compute weighted value
        weighted_value = (weight_year * year_avg +
                          weight_region * region_avg +
                          weight_state * state_avg)

        # Debug: Print the computed weighted value
        print(f"Filling cell at row {row}, col {col} with {weighted_value:.2f} "
              f"(YearAvg: {year_avg}, StateAvg: {state_avg}, RegionAvg: {region_avg})")

        # Fill the empty cell
        cell.value = weighted_value

# Save the updated file
wb.save("file_name.xlsx")
print("Your file is saved in the project folder")
