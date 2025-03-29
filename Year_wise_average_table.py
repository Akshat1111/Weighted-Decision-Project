from openpyxl import load_workbook
# Calculate and make a table for year wise average in excel, transposing the current dataset
# Load the workbook
wb = load_workbook("prices.xlsx")  # replace with your filename
data_sheet = wb["Data"]
summary_sheet = wb["Sheet1"]

# Prepare a list to store (year, average) tuples
results = []

# Loop through the columns E to L (Excel columns E=5 to L=12)
for col in range(5, 13):
    # Get the year header from row 4 of the current column
    year = data_sheet.cell(row=4, column=col).value
    
    # Collect numeric values from row 5 to 167 (ignore header in row 4)
    values = []
    for row in range(5, 168):
        cell_value = data_sheet.cell(row=row, column=col).value
        if isinstance(cell_value, (int, float)):
            values.append(cell_value)
    
    # Calculate the average if there are any values
    if values:
        avg = sum(values) / len(values)
    else:
        avg = None  # or use 0, or leave it blank as needed
    
    results.append((year, avg))

# Write headers in "sheet1": I4 for Year and J4 for Average
summary_sheet["I4"] = "Year"
summary_sheet["J4"] = "Average"

# Write the computed data starting at row 5
current_row = 5
for year, avg in results:
    summary_sheet.cell(row=current_row, column=9, value=year)  # Column I is 9
    summary_sheet.cell(row=current_row, column=10, value=avg)    # Column J is 10
    current_row += 1

# Save the workbook under a new name (or overwrite)
wb.save("example_updated.xlsx")
