from openpyxl import load_workbook

# Load the workbook
wb = load_workbook("example_updated_6.xlsx")  # Replace with actual filename
data_sheet = wb["Data(2)"]
sheet1 = wb["Sheet1"]

# Step 1: Extract precomputed data from "Sheet1"
# Years mapping from Sheet1 I5:I12 (year labels) → J5:J12 (corresponding avg values)
year_avg_map = {sheet1.cell(row=i, column=9).value: sheet1.cell(row=i, column=10).value for i in range(5, 13)}

# State-level mapping from M5:O72 (State, Occurrence, Average)
state_avg_map = {sheet1.cell(row=i, column=13).value: (sheet1.cell(row=i, column=14).value, sheet1.cell(row=i, column=15).value) for i in range(5, 73)}

# Region-level mapping from Q5:S8 (Region, Occurrence, Average)
region_avg_map = {sheet1.cell(row=i, column=17).value: (sheet1.cell(row=i, column=18).value, sheet1.cell(row=i, column=19).value) for i in range(5, 9)}

# Step 2: Process each column in "Data" for years (E5:L167)
for col in range(5, 13):  # E = 5, L = 12
    year = data_sheet.cell(row=4, column=col).value  # Year header

    if year not in year_avg_map:
        continue  # Skip if the year is not found

    year_avg = year_avg_map[year]

    for row in range(5, 168):  # E5:L167
        cell = data_sheet.cell(row=row, column=col)

        if cell.value not in [None, ""]:
            continue  # Skip non-empty cells

        # Get state and region for this row
        state = data_sheet.cell(row=row, column=3).value  # Column C
        region = data_sheet.cell(row=row, column=4).value  # Column D

        # Get precomputed state & region averages
        state_data = state_avg_map.get(state, (0, year_avg))  # (Occurrence, Average)
        region_data = region_avg_map.get(region, (0, year_avg))  # (Occurrence, Average)

        state_occurrence, state_avg = state_data
        region_occurrence, region_avg = region_data

        # Determine weights based on state occurrence
        if state_occurrence < 3:
            weight_state = 0.0
            weight_region = 0.5
        else:
            weight_state = 0.2
            weight_region = 0.3
        weight_year = 0.5

        # Compute weighted value
        weighted_value = (weight_year * year_avg +
                          weight_region * region_avg +
                          weight_state * state_avg)

        # Fill the empty cell
        cell.value = weighted_value

# Save the updated file
wb.save("example_updated_7.xlsx")
