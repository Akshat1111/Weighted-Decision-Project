from openpyxl import load_workbook

# Load the workbook and sheets
wb = load_workbook("example_updated_2.xlsx")  # Replace with your filename
data_sheet = wb["Data"]
summary_sheet = wb["Sheet1"]

# Dictionary to store region information:
# Key: region string, Value: dict with 'count' and 'values' list.
region_data = {}

# Loop over rows 5 to 167 in "Data" (skip header row 4)
for row in range(5, 168):
    region = data_sheet.cell(row=row, column=4).value  # Column D is index 4
    if region is None or str(region).strip() == "":
        continue  # Skip rows with empty region
    # Initialize dictionary entry if not already present
    if region not in region_data:
        region_data[region] = {'count': 0, 'values': []}
    
    # Increase the occurrence count for the region
    region_data[region]['count'] += 1

    # Loop over columns E to L (columns 5 to 12) for the current row
    for col in range(5, 13):
        cell_value = data_sheet.cell(row=row, column=col).value
        # Only add numeric values (skip blanks or non-numbers)
        if isinstance(cell_value, (int, float)):
            region_data[region]['values'].append(cell_value)

# Create a sorted list of regions for consistent output
sorted_regions = sorted(region_data.keys())

# Write the headers in "sheet1" starting at Q4, R4, and S4
summary_sheet["Q4"] = "Region"
summary_sheet["R4"] = "Occurrence"
summary_sheet["S4"] = "Average"

# Starting row for the summary table on "sheet1" (from row 5 downward)
start_row = 5

# Write each region's data to the summary table
for idx, region in enumerate(sorted_regions):
    row = start_row + idx
    count = region_data[region]['count']
    values = region_data[region]['values']
    # Calculate the average if there are any values; else use None (or zero)
    avg = sum(values) / len(values) if values else None

    summary_sheet.cell(row=row, column=17, value=region)  # Column Q is index 17
    summary_sheet.cell(row=row, column=18, value=count)     # Column R is index 18
    summary_sheet.cell(row=row, column=19, value=avg)         # Column S is index 19

# Save the updated workbook (change the filename as needed)
wb.save("example_updated_3.xlsx")
